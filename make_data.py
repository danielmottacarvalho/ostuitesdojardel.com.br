import openpyxl
import json
import re
from datetime import datetime
from collections import Counter

XLSX_PATH = '/Users/danielmotta/Documents/jardel-memorial/Jardel Final.xlsx'
OUTPUT_PATH = '/Users/danielmotta/Documents/jardel-memorial/data.js'

# ── Keyword lists ─────────────────────────────────────────────────────────────
# Rules:
#  • Every keyword is matched as a whole word/phrase using \b boundaries.
#  • Include both accented and unaccented forms where tweets may omit accents.
#  • No need for space-padding tricks — regex handles boundaries cleanly.

CATEGORIES = [
    ('Futebol', [
        'futebol', 'gol', 'goleiro', 'jogador', 'clube', 'campeonato', 'copa',
        'liga', 'corinthians', 'palmeiras', 'flamengo', 'grêmio', 'gremia',
        'cruzeiro', 'atlético', 'atletico', 'vasco', 'botafogo', 'fluminense',
        'torcida', 'estádio', 'estadio', 'técnico', 'tecnico', 'placar',
        'partida', 'rodada', 'brasileiro', 'bola', 'zagueiro', 'lateral',
        'atacante', 'volante', 'árbitro', 'arbitro', 'pênalti', 'penalti',
        'escanteio', 'são paulo fc', 'spfc', 'tricolor', 'seleção', 'selecao',
        'fifa', 'uefa', 'premier league', 'libertadores', 'sulamericana',
        'mundial', 'cbf', 'torcedor', 'beisebol', 'basquete', 'nba', 'nfl',
        'tênis', 'esporte', 'time', 'jogo',
    ]),
    ('Música', [
        'música', 'musica', 'canção', 'cancao', 'cantor', 'cantora', 'banda',
        'álbum', 'album', 'discografia', 'show', 'rock', 'jazz', 'samba',
        'forró', 'forro', 'rap', 'hip-hop', 'festival', 'palco', 'disco',
        'melodia', 'ritmo', 'playlist', 'spotify', 'vinil', 'guitarra',
        'piano', 'bateria', 'musical', 'concerto', 'ópera', 'opera', 'mpb',
        'axé', 'axe', 'pagode', 'funk', 'eletrônica', 'eletronica', 'dj',
        'trilha sonora', 'refrão', 'refrao', 'single', 'riff', 'acústico',
        'acustico', 'vocal', 'punk', 'metal', 'clássico', 'classico',
        'groove', 'reggae', 'bossa', 'bolero', 'balada', 'faixa', 'mixtape',
        'lp', 'ep', 'ouço', 'ouvi', 'ouvindo', 'tocando', 'cantando',
        'turnê', 'turne', 'gravadora', 'clipe', 'lyric', 'lyrics',
        'gogol bordello', 'beatles', 'rolling stones', 'radiohead', 'nirvana',
        'manu chao', 'caetano', 'veloso', 'chico buarque', 'cazuza',
        'renato russo', 'roberto carlos', 'gal costa', 'maria bethânia',
        'maria bethania', 'letra', 'voz',
    ]),
    ('Televisão', [
        'televisão', 'televisao', 'tv', 'canal', 'emissora', 'programa',
        'série', 'serie', 'novela', 'globo', 'sbt', 'record', 'band',
        'telenovela', 'ator', 'atriz', 'episódio', 'episodio', 'temporada',
        'netflix', 'amazon prime', 'hbo', 'disney+', 'personagem', 'elenco',
        'roteiro', 'transmissão', 'transmissao', 'telejornal', 'reality',
        'big brother', 'bbb', 'documentário', 'documentario', 'minissérie',
        'globoplay', 'telecine', 'assistindo', 'assisti', 'season', 'finale',
        'bial', 'gugu', 'superpop',
    ]),
    ('Jornalismo', [
        'jornalismo', 'jornalista', 'reportagem', 'notícia', 'noticia',
        'jornal', 'revista', 'mídia', 'midia', 'imprensa', 'redação',
        'redacao', 'editor', 'pauta', 'entrevista', 'coluna', 'artigo',
        'portal', 'manchete', 'cobertura', 'correspondente', 'matéria',
        'materia', 'editorial', 'repórter', 'reporter', 'agência', 'agencia',
        'reuters', 'folha', 'estadão', 'estadao', 'veja', 'época', 'epoca',
        'piauí', 'piaui', 'carta capital', 'newsletter', 'scoop',
        'off the record', 'apuração', 'apuracao', 'checagem', 'fact-check',
        'desinformação', 'desinformacao', 'fake news', 'blog',
    ]),
    ('Política', [
        'política', 'politica', 'político', 'politico', 'presidente',
        'governo', 'congresso', 'senado', 'câmara', 'camara', 'eleição',
        'eleicao', 'partido', 'candidato', 'voto', 'democracia', 'ditadura',
        'corrupção', 'corrupcao', 'pt', 'psdb', 'pmdb', 'lula', 'dilma',
        'bolsonaro', 'temer', 'ministro', 'ministério', 'ministerio',
        'reforma', 'lei', 'constituição', 'constituicao', 'stf', 'deputado',
        'senador', 'governador', 'prefeito', 'república', 'republica',
        'oposição', 'oposicao', 'orçamento', 'orcamento', 'imposto',
        'privatização', 'privatizacao', 'esquerda', 'direita', 'ideologia',
        'populismo', 'neoliberal', 'socialismo', 'capitalismo', 'manifestação',
        'manifestacao', 'protesto', 'greve', 'impeachment', 'cpi',
        'lava jato', 'procurador', 'delação', 'delacao', 'haddad', 'ciro',
        'alckmin', 'aécio', 'aecio', 'doria', 'moro', 'poder', 'federal',
        'trump', 'urnas', 'fascista', 'fascismo', 'fascist', 'candidatura',
    ]),
    ('São Paulo', [
        'são paulo', 'sao paulo', 'sampa', 'paulista', 'paulistano',
        'paulistana', 'avenida paulista', 'metrô', 'metro', 'pinheiros',
        'jardins', 'moema', 'itaim', 'vila madalena', 'higienópolis',
        'higienopolis', 'consolação', 'consolacao', 'berrini', 'faria lima',
        'paulicéia', 'pauliceia', 'liberdade', 'bela vista', 'ibirapuera',
        'ipiranga', 'morumbi', 'brooklin', 'aclimação', 'aclimacao',
        'bom retiro', 'brás', 'lapa', 'perdizes', 'pompeia', 'sp',
    ]),
    ('Pessoas', [
        'meu amigo', 'minha amiga', 'feliz aniversário', 'feliz aniversario',
        'parabéns', 'parabens', 'saudade', 'homenagem', 'descanse em paz',
        'faleceu', 'morreu', 'nasceu', 'casou', 'noiva', 'noivo',
        'meu pai', 'minha mãe', 'minha mae', 'meu filho', 'minha filha',
        'meu irmão', 'minha irmã', 'grande amigo', 'melhor amigo',
        'meu brother', 'obrigado a', 'obrigada a',
    ]),
]

# Precompile one pattern per keyword: \b<keyword>\b, case-insensitive
COMPILED = [
    (cat, [re.compile(r'\b' + re.escape(kw) + r'\b', re.IGNORECASE | re.UNICODE)
           for kw in keywords])
    for cat, keywords in CATEGORIES
]

def categorize(text):
    if not text:
        return 'Outros'
    for category, patterns in COMPILED:
        for pat in patterns:
            if pat.search(text):
                return category
    return 'Outros'

# ── Read & filter ─────────────────────────────────────────────────────────────
print("Reading Excel file...")
wb = openpyxl.load_workbook(XLSX_PATH, read_only=True)
ws = wb.active
headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
col = {name: i for i, name in enumerate(headers)}

tweets = []
for row in ws.iter_rows(min_row=2, values_only=True):
    row = row + (None,) * (len(headers) - len(row))
    if row[col['User ID']] != 'jsebba':
        continue

    text       = row[col['Tweet Text']] or ''
    created    = row[col['Created On']]
    bookmarks  = row[col['Bookmark #']] or 0
    likes      = row[col['Favorite #']] or 0
    views      = row[col['View #']] or 0
    media_url  = row[col['Media URL']] or None
    tweet_type = row[col['Type']] or ''

    date_str = created.strftime('%Y-%m-%dT%H:%M:%S') if isinstance(created, datetime) else (str(created) if created else '')

    tweets.append({
        'text':      text,
        'date':      date_str,
        'likes':     int(likes),
        'views':     int(views) if views else 0,
        'bookmarks': int(bookmarks),
        'type':      tweet_type,
        'mediaUrl':  media_url,
        'category':  categorize(text),
    })

wb.close()
print(f"Filtered to {len(tweets)} jsebba tweets.")

# ── Category breakdown ────────────────────────────────────────────────────────
counts = Counter(t['category'] for t in tweets)
print("\nCategory breakdown:")
for cat, n in sorted(counts.items(), key=lambda x: -x[1]):
    print(f"  {cat:<15} {n:>5}  ({n/len(tweets)*100:.1f}%)")

# ── Spot-check known false-positive cases ─────────────────────────────────────
print("\nSpot-checks (should NOT be Futebol):")
tests = [
    "jogaria a pessoa na primeira fila de um show do Gogol Bordello",
    "Falaram hoje em Brett Anderson no Brasil em janeiro",
    "Maravilhoso espetáculo no teatro ontem",
]
for t in tests:
    print(f"  [{categorize(t)}] {t[:70]}")

# ── Write data.js ─────────────────────────────────────────────────────────────
print(f"\nWriting {OUTPUT_PATH}...")
json_str = json.dumps(tweets, ensure_ascii=False, indent=2)
with open(OUTPUT_PATH, 'w', encoding='utf-8') as f:
    f.write('const tweetData = ')
    f.write(json_str)
    f.write(';\n')

print("Done!")
