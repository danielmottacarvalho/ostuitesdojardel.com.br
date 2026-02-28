import openpyxl
import urllib.request
import os
from urllib.parse import urlparse
from concurrent.futures import ThreadPoolExecutor, as_completed
import threading

XLSX_PATH = '/Users/danielmotta/Documents/jardel-memorial/Jardel Final.xlsx'
IMAGES_DIR = '/Users/danielmotta/Documents/jardel-memorial/images'

# Step 1: collect all pbs.twimg.com URLs for jsebba
print("Reading Excel file...")
wb = openpyxl.load_workbook(XLSX_PATH, read_only=True)
ws = wb.active
headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
user_id_col = headers.index('User ID')
media_col = headers.index('Media URL')

rows_with_media = []  # (row_index, url)
for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
    row = row + (None,) * (len(headers) - len(row))
    if row[user_id_col] == 'jsebba' and row[media_col] and 'pbs.twimg.com' in str(row[media_col]):
        rows_with_media.append((i, row[media_col]))
wb.close()

print(f"Found {len(rows_with_media)} images to download.\n")

# Step 2: download with thread pool
lock = threading.Lock()
done = 0
failed = []

def download(item):
    row_idx, url = item
    filename = os.path.basename(urlparse(url).path)
    dest = os.path.join(IMAGES_DIR, filename)
    if os.path.exists(dest):
        return (row_idx, url, filename, True, 'already exists')
    try:
        req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
        with urllib.request.urlopen(req, timeout=30) as r:
            data = r.read()
        with open(dest, 'wb') as f:
            f.write(data)
        return (row_idx, url, filename, True, f'{len(data):,} bytes')
    except Exception as e:
        return (row_idx, url, None, False, str(e))

print("Downloading images (20 concurrent)...")
results = {}
with ThreadPoolExecutor(max_workers=20) as executor:
    futures = {executor.submit(download, item): item for item in rows_with_media}
    for future in as_completed(futures):
        row_idx, url, filename, ok, msg = future.result()
        with lock:
            done += 1
            if ok:
                results[row_idx] = filename
            else:
                failed.append((url, msg))
            if done % 100 == 0 or done == len(rows_with_media):
                print(f"  {done}/{len(rows_with_media)} done, {len(failed)} failed")

print(f"\nDownload complete. {len(results)} succeeded, {len(failed)} failed.")
if failed:
    print("Failed URLs:")
    for url, err in failed:
        print(f"  {url}: {err}")

# Step 3: update Excel with local paths
print("\nUpdating Excel file with local image paths...")
wb2 = openpyxl.load_workbook(XLSX_PATH)
ws2 = wb2.active
headers2 = [cell.value for cell in ws2[1]]
media_col2 = headers2.index('Media URL') + 1  # 1-indexed

updated = 0
for row_idx, filename in results.items():
    local_path = f'images/{filename}'
    ws2.cell(row=row_idx, column=media_col2).value = local_path
    updated += 1

wb2.save(XLSX_PATH)
print(f"Updated {updated} rows in Excel file.")
print("Done!")
